"""File watcher — monitors directories for inventory, sales, and work files.

Bridges watchdog (sync/threaded) to Conduit's async notification pipeline.
Ported from spectre/scripts/downloads_watcher.py with async notifications.
"""

import asyncio
import logging
import os
import re
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

from . import config, db, ntfy, spectre
from . import telegram as tg_module
from .ws import ConnectionManager

log = logging.getLogger("conduit.watcher")

# Extensions we care about
EXCEL_EXTENSIONS = {".xlsx", ".xls"}
PDF_EXTENSION = ".pdf"
WORK_EXTENSIONS = EXCEL_EXTENSIONS | {PDF_EXTENSION, ".csv"}

# Sort category keywords
SORT_CATEGORIES = {
    "Invoices": ["invoice", "inv_", "inv-", "billing", "charge"],
    "Receipts": ["receipt", "rcpt", "payment", "confirmation"],
    "Training": ["training", "manual", "guide", "tutorial", "instruction", "sop", "procedure"],
    "Reports": ["report", "summary", "analysis", "audit", "review"],
    "Contracts": ["contract", "agreement", "terms", "nda"],
    "Other": ["order", "purchase", "vendor", "inventory", "menu", "schedule", "roster"],
}

# Sales PDF filename pattern (e.g. "Financial Detail Report" or date-based names)
SALES_PDF_PATTERN = re.compile(
    r"(financial.?detail|daily.?sales|pos.?report|sales.?report)", re.IGNORECASE
)

# Module state
_manager: ConnectionManager | None = None
_loop: asyncio.AbstractEventLoop | None = None


def _sort_base() -> Path:
    return Path(os.path.expanduser(config.WATCHER_SORT_BASE))


# ---------------------------------------------------------------------------
# Detection helpers (ported from spectre downloads_watcher.py)
# ---------------------------------------------------------------------------

def is_spectre_inventory(filepath: Path) -> tuple[bool, dict[str, Any] | None]:
    """Check if an Excel file is a Spectre inventory by content.

    Looks for date in row 1 and site name in row 2.
    """
    if filepath.suffix.lower() not in EXCEL_EXTENSIONS:
        return False, None

    try:
        from openpyxl import load_workbook
    except ImportError:
        log.debug("openpyxl not available — skipping inventory detection")
        return False, None

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)

        # Find data sheet
        data_sheet = None
        for name in wb.sheetnames:
            if "data" in name.lower():
                data_sheet = wb[name]
                break
        if not data_sheet:
            for name in wb.sheetnames:
                lower = name.lower()
                if "inventory" in lower and "summary" not in lower:
                    data_sheet = wb[name]
                    break
        if not data_sheet:
            data_sheet = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else wb.active

        if not data_sheet:
            wb.close()
            return False, None

        # Row 1: date (MM/DD/YYYY)
        row1_val = data_sheet.cell(row=1, column=1).value
        date_match = None
        if row1_val and isinstance(row1_val, str):
            date_match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})$", row1_val.strip())

        # Row 2: site name (text before first parenthesis)
        row2_val = data_sheet.cell(row=2, column=1).value
        site_name = None
        if row2_val and isinstance(row2_val, str):
            site_match = re.match(r"^([^(]+)", row2_val.strip())
            if site_match:
                site_name = site_match.group(1).strip()

        wb.close()

        if not date_match or not site_name:
            return False, None

        month, day, year = date_match.groups()
        metadata = {
            "inventory_date": f"{year}-{month.zfill(2)}-{day.zfill(2)}",
            "site_name": site_name,
            "site_id": re.sub(r"[^a-z0-9]+", "_", site_name.lower()).strip("_"),
        }
        return True, metadata

    except Exception as e:
        log.debug("Error checking Excel content: %s", e)
        return False, None


def categorize_file(filepath: Path) -> str | None:
    """Categorize a file by filename keywords. Returns folder name or None."""
    filename = filepath.name.lower()

    for category, keywords in SORT_CATEGORIES.items():
        if any(kw in filename for kw in keywords):
            return category

    return None


def is_sales_pdf(filepath: Path) -> bool:
    """Check if a PDF looks like a sales/financial report."""
    if filepath.suffix.lower() != PDF_EXTENSION:
        return False
    return bool(SALES_PDF_PATTERN.search(filepath.name))


def parse_sales_pdf(filepath: Path) -> dict[str, Any] | None:
    """Extract revenue and cover counts from a sales PDF using pdfplumber."""
    try:
        import pdfplumber
    except ImportError:
        log.debug("pdfplumber not available — skipping sales parsing")
        return None

    try:
        with pdfplumber.open(filepath) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)

        if not text.strip():
            return None

        result: dict[str, Any] = {"filename": filepath.name}

        # Extract dollar amounts — look for net sales patterns
        # Patterns like "Net Sales $4,821.50" or "NET SALES 4821.50"
        net_pattern = re.compile(
            r"net\s*sales[:\s]*\$?([\d,]+\.?\d*)", re.IGNORECASE
        )
        net_match = net_pattern.search(text)
        if net_match:
            result["net_sales"] = float(net_match.group(1).replace(",", ""))

        # Fallback: largest dollar amount on the page
        if "net_sales" not in result:
            amounts = re.findall(r"\$?([\d,]+\.\d{2})", text)
            if amounts:
                values = [float(a.replace(",", "")) for a in amounts]
                # Filter to reasonable range ($100 - $50,000)
                reasonable = [v for v in values if 100 <= v <= 50000]
                if reasonable:
                    result["net_sales"] = max(reasonable)

        # Extract cover/guest count
        cover_pattern = re.compile(
            r"(?:covers?|guests?|checks?)\s*[:\s]*(\d+)", re.IGNORECASE
        )
        cover_match = cover_pattern.search(text)
        if cover_match:
            result["covers"] = int(cover_match.group(1))

        # Extract date from content
        date_pattern = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")
        date_match = date_pattern.search(text)
        if date_match:
            m, d, y = date_match.groups()
            if len(y) == 2:
                y = f"20{y}"
            result["report_date"] = f"{y}-{m.zfill(2)}-{d.zfill(2)}"

        return result if len(result) > 1 else None  # Must have more than just filename

    except Exception as e:
        log.error("Sales PDF parse error: %s", e)
        return None


# ---------------------------------------------------------------------------
# Async file processing (called from watchdog thread via run_coroutine_threadsafe)
# ---------------------------------------------------------------------------

async def _notify_all(title: str, body: str, tags: list[str] | None = None, priority: int = 3):
    """Push notification to WS, ntfy, and Telegram."""
    if _manager:
        await _manager.push(content=f"**{title}**\n{body}", title=title)
    await ntfy.push(title=title, body=body, tags=tags or ["file_folder"], priority=priority)
    await tg_module.push(title=title, body=body)


async def process_file(filepath: Path, actions: list[str]):
    """Async orchestrator: classify → act → notify all channels."""
    if not filepath.exists():
        return

    ext = filepath.suffix.lower()

    # Skip non-work files
    if ext not in WORK_EXTENSIONS:
        return

    log.info("Processing: %s", filepath.name)

    # 1. Check for inventory xlsx
    if "detect_inventory" in actions and ext in EXCEL_EXTENSIONS:
        is_inv, metadata = is_spectre_inventory(filepath)
        if is_inv and metadata:
            site = metadata.get("site_name", "unknown")
            date = metadata.get("inventory_date", "unknown")
            site_id = metadata.get("site_id")

            # Upload to Spectre
            success, is_dup, _ = await spectre.upload_file(filepath, site_id)
            if success:
                filepath.unlink(missing_ok=True)
                status = "already in Spectre" if is_dup else "uploaded to Spectre"
                await _notify_all(
                    "Inventory Uploaded",
                    f"Inventory for {site} ({date}) {status}.",
                    tags=["package"],
                )
            else:
                await _notify_all(
                    "Inventory Upload Failed",
                    f"Could not upload {filepath.name} to Spectre. Is it running?",
                    tags=["warning"],
                    priority=4,
                )
            return

    # 2. Check for sales PDF
    if "detect_sales" in actions or "parse_sales" in actions:
        if is_sales_pdf(filepath):
            data = parse_sales_pdf(filepath)
            if data:
                parts = []
                if "net_sales" in data:
                    parts.append(f"${data['net_sales']:,.2f} net")
                if "covers" in data:
                    parts.append(f"{data['covers']} covers")
                if "report_date" in data:
                    try:
                        dt = datetime.strptime(data["report_date"], "%Y-%m-%d")
                        parts.append(dt.strftime("%a %m/%d"))
                    except ValueError:
                        pass
                summary = ", ".join(parts) if parts else filepath.name
                await _notify_all(
                    "Sales Report",
                    f"Sales report: {summary}",
                    tags=["chart_with_upwards_trend"],
                )

                # If this was in Downloads, move to sales folder
                if "detect_sales" in actions:
                    sales_dir = Path(os.path.expanduser("~/Documents/Work/lockheed/sales/"))
                    sales_dir.mkdir(parents=True, exist_ok=True)
                    dest = sales_dir / filepath.name
                    if not dest.exists():
                        shutil.move(str(filepath), str(dest))
                        log.info("Moved sales PDF to %s", dest)
                return

    # 3. Sort work files
    if "sort" in actions:
        category = categorize_file(filepath)
        if category:
            dest_dir = _sort_base() / category
            dest_dir.mkdir(parents=True, exist_ok=True)
            dest = dest_dir / filepath.name

            # Handle duplicate names
            if dest.exists():
                base = filepath.stem
                ext_str = filepath.suffix
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest = dest_dir / f"{base}_{ts}{ext_str}"

            shutil.move(str(filepath), str(dest))
            log.info("Sorted %s -> %s/", filepath.name, category)
            await _notify_all(
                "File Sorted",
                f"Sorted {filepath.name} -> {category}/",
                tags=["file_folder"],
                priority=2,
            )


# ---------------------------------------------------------------------------
# Watchdog handler (sync thread → async bridge)
# ---------------------------------------------------------------------------

class ConduitFileHandler(FileSystemEventHandler):
    """Bridges watchdog events to Conduit's async pipeline."""

    def __init__(self, actions: list[str], debounce: float = 3.0):
        self.actions = actions
        self.debounce = debounce
        self._recent: dict[str, float] = {}

    def _should_process(self, path: str) -> bool:
        now = time.time()
        if path in self._recent and now - self._recent[path] < self.debounce:
            return False
        self._recent[path] = now
        # Prune old entries
        self._recent = {k: v for k, v in self._recent.items() if now - v < 60}
        return True

    def _dispatch(self, filepath: Path):
        if not _loop:
            return
        # Wait for file to finish writing
        time.sleep(config.WATCHER_DEBOUNCE)
        if not filepath.exists():
            return
        asyncio.run_coroutine_threadsafe(process_file(filepath, self.actions), _loop)

    def on_created(self, event):
        if event.is_directory:
            return
        if not self._should_process(event.src_path):
            return
        self._dispatch(Path(event.src_path))

    def on_moved(self, event):
        if event.is_directory:
            return
        if not self._should_process(event.dest_path):
            return
        self._dispatch(Path(event.dest_path))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def start(manager: ConnectionManager) -> Observer | None:
    """Start file watching. Called from app.py lifespan."""
    global _manager, _loop

    if not config.WATCHER_ENABLED:
        return None

    _manager = manager
    _loop = asyncio.get_running_loop()

    observer = Observer()
    observer.daemon = True

    watched = 0
    for dir_cfg in config.WATCHER_DIRECTORIES:
        path = os.path.expanduser(dir_cfg.get("path", ""))
        actions = dir_cfg.get("actions", [])
        if not path or not os.path.isdir(path):
            log.warning("Watcher: skipping non-existent directory %s", path)
            continue

        handler = ConduitFileHandler(actions=actions, debounce=config.WATCHER_DEBOUNCE)
        observer.schedule(handler, path, recursive=False)
        watched += 1
        log.info("Watching: %s (actions: %s)", path, actions)

    if watched == 0:
        log.warning("Watcher enabled but no valid directories configured")
        return None

    # Ensure sort directories exist
    sort_base = _sort_base()
    for category in SORT_CATEGORIES:
        (sort_base / category).mkdir(parents=True, exist_ok=True)

    observer.start()
    log.info("File watcher started (%d directories)", watched)
    return observer


def stop(observer: Observer | None):
    """Stop the file watcher. Called from app.py shutdown."""
    if observer:
        observer.stop()
        observer.join(timeout=5)
        log.info("File watcher stopped")
